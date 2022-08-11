import xlsxwriter

from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl.drawing.image import Image

from openpyxl.styles.borders import Border, Side

from openpyxl.styles import Font, Alignment,PatternFill

import numpy as np

def simpan_excel(pool_path_simpan,nama_file_analitik,pool_json_raw,timeType,keluar_analitik):
    
    workbook = xlsxwriter.Workbook('{}\{} per {}.xlsx'.format(pool_path_simpan,nama_file_analitik,timeType))

    Analytics = workbook.add_worksheet('Analytics')
    Analytics.write('P2', 'Kec_angin')
    Analytics.write('Q2', 'Max')
    Analytics.write('Q3', 'Mean')
    Analytics.write('Q4', 'Std')
    Analytics.write('R1', 'U')
    Analytics.write('S1', 'TL')
    Analytics.write('T1', 'T')
    Analytics.write('U1', 'Tg')
    Analytics.write('V1', 'S')
    Analytics.write('W1', 'BD')
    Analytics.write('X1', 'B')
    Analytics.write('Y1', 'BL')
    Analytics.write('P6', 'Theoritical Power')
    Analytics.write('Q6', 'Max')
    Analytics.write('Q7', 'Mean')
    Analytics.write('Q8', 'Std')




    RAW_Data = workbook.add_worksheet('RAW_Data')
    RAW_Data.write('A1', 'ID_Device')
    RAW_Data.write('B1', 'Date')
    RAW_Data.write('C1', 'Time')
    RAW_Data.write('D1', 'Kec_angin')
    RAW_Data.write('E1', 'Arah_angin')
    RAW_Data.write('F1', 'Derajat_angin')
    RAW_Data.write('G1', 'Power')


    row = 1
    col = 0

    max_kec=keluar_analitik[0]
    max_power=keluar_analitik[1]
    mean_kec=keluar_analitik[2]
    mean_power=keluar_analitik[3]
    std_kec=keluar_analitik[4]
    std_power=keluar_analitik[5]

    counter_analitik=0
    for A_proses in max_kec:
        Analytics.write(row,col+17+counter_analitik,max_kec[counter_analitik])
        Analytics.write(row+1,col+17+counter_analitik,mean_kec[counter_analitik])
        Analytics.write(row+2,col+17+counter_analitik,std_kec[counter_analitik])

        Analytics.write(row+4,col+17+counter_analitik,max_power[counter_analitik])
        Analytics.write(row+5,col+17+counter_analitik,mean_power[counter_analitik])
        Analytics.write(row+6,col+17+counter_analitik,std_power[counter_analitik])


        
        if (counter_analitik==max_kec[-1]):
            counter_analitik=0
        else:
            counter_analitik+=1
    cell_format = workbook.add_format()
    cell_format.set_num_format('hh:mm:ss')

    for raw in (pool_json_raw):
        # Convert the date string into a datetime object.
        # time_strp = datetime.strptime(time_detik, "%H:%M:%S")


        date_time = datetime.strptime(raw['Time'], '%H:%M:%S')

        RAW_Data.write(row, col, raw['ID_Device'])
        RAW_Data.write(row, col+1, raw['Date'] )
        RAW_Data.write_datetime(row, col+2, date_time,cell_format )
        RAW_Data.write_number(row, col+3, raw['Kec_angin'] )
        RAW_Data.write(row, col+4, raw['Arah_angin'] )
        RAW_Data.write(row, col+5, raw['Derajat_angin'] )
        RAW_Data.write_number(row, col+6, raw['Power'] )

        if (raw==pool_json_raw[-1]):
            row=1
        else:
            row += 1

    if (timeType=="detik"):
        range=86401
        interval=10801
    elif (timeType=="menit"):
        range=1441
        interval=360

    # Create a chart object.
    chart = workbook.add_chart({'type': 'scatter','subtype':'straight'})
    chart.add_series({'categories': '=RAW_Data!$C$1:$C${}'.format(range),'values':'=RAW_Data!$D$1:$D${}'.format(range)})
    chart.set_legend({'delete_series': [0]})
    chart.set_x_axis({'date_axis': True,'interval_unit': interval,'num_font':  {'rotation': 45}})
    chart.set_y_axis({'name':'Kecepatan Angin (m/s)'})
    chart.set_size({'width': 800, 'height': 300})
    Analytics.insert_chart(1, 1, chart)


    workbook.close()


def cari_count(json_pool,kec_low,kec_high,kec_0_2):

    arah_angin=["U","TL","T","Tg","S","BD","B","BL"]
    


    jumlah=0
    for i in arah_angin:
        if kec_low!=10:
            coba = sum(((x['Arah_angin']==i) and ((x['Kec_angin']<kec_high) and (x['Kec_angin']>=kec_low)) ) for x in json_pool)
            
        else:
            coba = sum(((x['Arah_angin']==i) and ((x['Kec_angin']>=kec_low)) ) for x in json_pool)
        jumlah+=coba
        kec_0_2.append(coba)
    kec_0_2.append(jumlah)
        



def simpan_excel_openpyxl(pool_path_simpan,nama_file_analitik,pool_json_raw,timeType,keluar_analitik,file_name):
    workbook = Workbook()

    border_normal = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    bold=Border(top=Side(style='thick'), 
                     bottom=Side(style='thick'),left=Side(style='thick'),right=Side(style='thick'))

    sheet1 = workbook.active
    sheet1.title = "Analytics"

    sheet2 = workbook.create_sheet("RAW_Data")
    sheet1.merge_cells('N3:N5')

    sheet1['N3']= 'Kec_angin (m/s)'
    sheet1['N3'].font = Font(bold=True)
    sheet1['N3'].alignment = Alignment(horizontal='center', vertical='center')


    sheet1['O3']= 'Max'
    sheet1['O3'].font = Font(bold=True)

    sheet1['O4']= 'Mean'
    sheet1['O4'].font = Font(bold=True)

    sheet1['O5']= 'Std'
    sheet1['O5'].font = Font(bold=True)

    sheet1['P2']= 'U'
    sheet1['P2'].font = Font(bold=True)

    sheet1['Q2']= 'TL'
    sheet1['Q2'].font = Font(bold=True)

    sheet1['R2']= 'T'
    sheet1['R2'].font = Font(bold=True)

    sheet1['S2']= 'Tg'
    sheet1['S2'].font = Font(bold=True)

    sheet1['T2']= 'S'
    sheet1['T2'].font = Font(bold=True)

    sheet1['U2']= 'BD'
    sheet1['U2'].font = Font(bold=True)

    sheet1['V2']= 'B'
    sheet1['V2'].font = Font(bold=True)

    sheet1['W2']= 'BL'
    sheet1['W2'].font = Font(bold=True)

    sheet1['X2']= 'ALL'
    sheet1['X2'].font = Font(bold=True)

    sheet1.merge_cells('N7:N9')
    sheet1['N7']= 'Theoritical Power (Watt)'
    sheet1['N7'].font = Font(bold=True)
    sheet1['N7'].alignment = Alignment(horizontal='center', vertical='center')
    
    sheet1['O7']= 'Max'
    sheet1['O7'].font = Font(bold=True)
    
    sheet1['O8']= 'Mean'
    sheet1['O8'].font = Font(bold=True)
    
    sheet1['O9']= 'Std'
    sheet1['O9'].font = Font(bold=True)

    a=np.arange(3,6)
    b=np.arange(15,25)
    for row_border in a:
        for col_border in b:
            sheet1.cell(row=row_border,column=col_border).border=border_normal
            sheet1.cell(row=row_border+4,column=col_border).border=border_normal

    a=6
    b=np.arange(14,25)
    for col_border in b:
        sheet1.cell(row=a,column=col_border).border=Border(top=Side(style='thick'), 
                     bottom=Side(style='thick'))

        if col_border==14:
            sheet1.cell(row=a,column=col_border).border=  Border(top=Side(style='thick'), 
                     bottom=Side(style='thick'),left=Side(style='thick'))
        elif col_border==24:
            sheet1.cell(row=a,column=col_border).border=  Border(top=Side(style='thick'), 
                     bottom=Side(style='thick'),right=Side(style='thick'))

        sheet1.cell(row=a,column=col_border).fill= PatternFill("solid", start_color="757575")
        #757575

    a=2
    b=np.arange(14,25)
    for col_border in b:
        sheet1.cell(row=a,column=col_border).border=Border(top=Side(style='thick'), 
                     bottom=Side(style='thick'),left=Side(style='thin'),right=Side(style='thin'))

        if col_border==14:
            sheet1.cell(row=a,column=col_border).border=  Border(top=Side(style='thick'), 
                     bottom=Side(style='thick'),left=Side(style='thick'),right=Side(style='thin'))
        elif col_border==24:
            sheet1.cell(row=a,column=col_border).border=  Border(top=Side(style='thick'), 
                     bottom=Side(style='thick'),right=Side(style='thick'),left=Side(style='thin'))
        #757575

    a=np.arange(3,6)
    b=14
    for row_border in a:
        sheet1.cell(row=row_border,column=b).border=Border(left=Side(style='thick'))
        sheet1.cell(row=row_border+4,column=b).border=Border(left=Side(style='thick'))

        sheet1.cell(row=row_border,column=b+10).border=Border(right=Side(style='thick'))
        sheet1.cell(row=row_border+4,column=b+10).border=Border(right=Side(style='thick'))

        if (row_border+4)==9:
            sheet1.cell(row=row_border+4,column=b).border=Border(left=Side(style='thick'),bottom=Side(style='thick'))
            
            sheet1.cell(row=row_border+4,column=b+10).border=Border(right=Side(style='thick'),bottom=Side(style='thick'))

    a=9
    b=np.arange(15,24)
    for col_border in b:
        sheet1.cell(row=a,column=col_border).border=Border(right=Side(style='thin'),left=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thick'))
    




    a=np.arange(19,28)
    b=np.arange(9,19)
    for row_border in a:
        for col_border in b:
            sheet1.cell(row=row_border,column=col_border).border=border_normal
    
    a=19
    b=np.arange(9,19)
    for col_border in b:
        sheet1.cell(row=a,column=col_border).border=Border(top=Side(style='thick'),bottom=Side(style='thick'))

        if (col_border==9):
            sheet1.cell(row=a,column=col_border).border=Border(top=Side(style='thick'),bottom=Side(style='thick'),left=Side(style='thick'))
        elif(col_border==18):
            sheet1.cell(row=a,column=col_border).border=Border(top=Side(style='thick'),bottom=Side(style='thick'),right=Side(style='thick'))
    

    a=np.arange(20,28)
    b=9
    for row_border in a:
        sheet1.cell(row=row_border,column=b).border=Border(left=Side(style='thick'),bottom=Side(style='thin'))
        sheet1.cell(row=row_border,column=b+9).border=Border(right=Side(style='thick'),bottom=Side(style='thin'))

        if (row_border==27):
            sheet1.cell(row=row_border,column=b).border=Border(bottom=Side(style='thick'),left=Side(style='thick'))
            sheet1.cell(row=row_border,column=b+9).border=Border(bottom=Side(style='thick'),right=Side(style='thick'))
    
    a=27
    b=np.arange(10,18)
    for col_border in b:
        sheet1.cell(row=a,column=col_border).border=Border(bottom=Side(style='thick'),right=Side(style='thin'),left=Side(style='thin'))







    
    sheet1['I21']= '0-2 m/s'
    sheet1['I21'].font = Font(bold=True)

    sheet1['I22']= '2-4 m/s'
    sheet1['I22'].font = Font(bold=True)

    sheet1['I23']= '4-6 m/s'
    sheet1['I23'].font = Font(bold=True)

    sheet1['I24']= '6-8 m/s'
    sheet1['I24'].font = Font(bold=True)

    sheet1['I25']= '8-10 m/s'
    sheet1['I25'].font = Font(bold=True)

    sheet1['I26']= '10> m/s'
    sheet1['I26'].font = Font(bold=True)

    sheet1['I27']= 'Total Keseluruhan'
    sheet1['I27'].font = Font(bold=True)

    sheet1.merge_cells('I19:R19')
    sheet1['I19']= 'Wind speed distribution(%)'
    sheet1['I19'].font = Font(bold=True)
    sheet1['I19'].alignment = Alignment(horizontal='center', vertical='center')

    sheet1['J20']= 'U'
    sheet1['J20'].font = Font(bold=True)

    sheet1['K20']= 'TL'
    sheet1['K20'].font = Font(bold=True)

    sheet1['L20']= 'T'
    sheet1['L20'].font = Font(bold=True)

    sheet1['M20']= 'Tg'
    sheet1['M20'].font = Font(bold=True)

    sheet1['N20']= 'S'
    sheet1['N20'].font = Font(bold=True)

    sheet1['O20']= 'BD'
    sheet1['O20'].font = Font(bold=True)

    sheet1['P20']= 'B'
    sheet1['P20'].font = Font(bold=True)

    sheet1['Q20']= 'BL'
    sheet1['Q20'].font = Font(bold=True)






    sheet2['A1']='ID_Device'
    sheet2['B1']='Date'
    sheet2['C1']='Time'
    sheet2['D1']='Kec_angin'
    sheet2['E1']='Arah_angin'
    sheet2['F1']='Derajat_angin'
    sheet2['G1']='Power'

    max_kec=keluar_analitik[0]
    max_power=keluar_analitik[1]
    mean_kec=keluar_analitik[2]
    mean_power=keluar_analitik[3]
    std_kec=keluar_analitik[4]
    std_power=keluar_analitik[5]



    iterate_row=2
    for raw in (pool_json_raw):
        date_time = datetime.strptime(raw['Time'], '%H:%M:%S').time()

        sheet2.cell(row=iterate_row, column=1).value = raw['ID_Device']
        sheet2.cell(row=iterate_row, column=2).value = raw['Date']
        sheet2.cell(row=iterate_row, column=3).value = date_time
        
        sheet2.cell(row=iterate_row, column=4).value = raw['Kec_angin']
        sheet2.cell(row=iterate_row, column=5).value = raw['Arah_angin']
        sheet2.cell(row=iterate_row, column=6).value = raw['Derajat_angin']
        sheet2.cell(row=iterate_row, column=7).value = raw['Power']

        if (raw==pool_json_raw[-1]):
            iterate_row=3
        else:
            iterate_row += 1
    counter_analitik=0
    for A_proses in max_kec:
        sheet1.cell(row=iterate_row,column=16+counter_analitik).value=(max_kec[counter_analitik])
        sheet1.cell(row=iterate_row,column=16+counter_analitik).number_format='0.00'
        sheet1.cell(row=iterate_row+1,column=16+counter_analitik).value=(mean_kec[counter_analitik])
        sheet1.cell(row=iterate_row+1,column=16+counter_analitik).number_format='0.00'
        sheet1.cell(row=iterate_row+2,column=16+counter_analitik).value=(std_kec[counter_analitik])
        sheet1.cell(row=iterate_row+2,column=16+counter_analitik).number_format='0.00'

        sheet1.cell(row=iterate_row+4,column=16+counter_analitik).value=(max_power[counter_analitik])
        sheet1.cell(row=iterate_row+4,column=16+counter_analitik).number_format='0.00'
        sheet1.cell(row=iterate_row+5,column=16+counter_analitik).value=(mean_power[counter_analitik])
        sheet1.cell(row=iterate_row+5,column=16+counter_analitik).number_format='0.00'
        sheet1.cell(row=iterate_row+6,column=16+counter_analitik).value=(std_power[counter_analitik])
        sheet1.cell(row=iterate_row+6,column=16+counter_analitik).number_format='0.00'


        if (counter_analitik==max_kec[-1]):
            counter_analitik=0
        else:
            counter_analitik+=1

    if (timeType=="detik"):
        range=86401
        interval=10801
    elif (timeType=="menit"):
        range=1441
        interval=360


    chart = ScatterChart()

    xvalues = Reference(sheet2, min_col = 3,
                    min_row = 2, max_row = range)
                     
    yvalues = Reference(sheet2, min_col = 4,
                    min_row = 2, max_row = range)

    # create a 1st series of data
    series = Series(values = yvalues, xvalues = xvalues, title ="kec angin")
    
    # add series data to the chart object
    chart.series.append(series)
    
        
    # set the title of the x-axis
    chart.x_axis.title = "Time"
    
    # set the title of the y-axis
    chart.y_axis.title = " Kec Angin (m/s)"

    chart.width=19.5
    
    # add chart to the sheet
    # the top-left corner of a chart
    # is anchored to cell E2 .
    sheet1.add_chart(chart, "A2")

    if timeType=="detik":
        img = Image('{}\Windrose bar\Per_Detik\Windrose bar {} per {}.png'.format(pool_path_simpan,file_name,timeType))
    else:
        img = Image('{}\Windrose bar\Per_Menit\Windrose bar {} per {}.png'.format(pool_path_simpan,file_name,timeType))  

    
    #adjusting size
    img.height=450
    img.width =450
    
    #adding img to cell A3

    sheet1.add_image(img, 'A20')

    kec_0_2=[]
    kec_2_4=[]
    kec_4_6=[]
    kec_6_8=[]
    kec_8_10=[]
    kec_10_inf=[]

    cari_count(pool_json_raw,0,2,kec_0_2)
    cari_count(pool_json_raw,2,4,kec_2_4)
    cari_count(pool_json_raw,4,6,kec_4_6)
    cari_count(pool_json_raw,6,8,kec_6_8)
    cari_count(pool_json_raw,8,10,kec_8_10)
    cari_count(pool_json_raw,10,0,kec_10_inf)

    pivot=[kec_0_2,kec_2_4,kec_4_6,kec_6_8,kec_8_10,kec_10_inf]

    row_pivot=21
    U=0.0
    TL=0.0
    T=0.0
    Tg=0.0
    S=0.0
    BD=0.0
    B=0.0
    BL=0.0
    for p in pivot:
        column_pivot=10
        for isi_pivot in p:
            sheet1.cell(row=row_pivot,column=column_pivot).value=((isi_pivot/len(pool_json_raw)))
            sheet1.cell(row=row_pivot,column=column_pivot).number_format = '0.0000%'
            column_pivot+=1
        U+=p[0]
        TL+=p[1]
        T+=p[2]
        Tg+=p[3]
        S+=p[4]
        BD+=p[5]
        B+=p[6]
        BL+=p[7]
        row_pivot+=1

    # sheet1['J27']= "{:.2%}".format((U/len(pool_json_raw)))
    # sheet1['K27']= "{:.2%}".format((TL/len(pool_json_raw)))
    # sheet1['L27']= "{:.2%}".format((T/len(pool_json_raw)))
    # sheet1['M27']= "{:.2%}".format((Tg/len(pool_json_raw)))
    # sheet1['N27']= "{:.2%}".format((S/len(pool_json_raw)))
    # sheet1['O27']= "{:.2%}".format((BD/len(pool_json_raw)))
    # sheet1['P27']= "{:.2%}".format((B/len(pool_json_raw)))
    # sheet1['Q27']= "{:.2%}".format((BL/len(pool_json_raw)))
    sheet1['J27']= (U/len(pool_json_raw))
    sheet1['J27'].number_format='0.0000%'
    sheet1['K27']= (TL/len(pool_json_raw))
    sheet1['K27'].number_format='0.0000%'
    sheet1['L27']= (T/len(pool_json_raw))
    sheet1['L27'].number_format='0.0000%'
    sheet1['M27']= (Tg/len(pool_json_raw))
    sheet1['M27'].number_format='0.0000%'
    sheet1['N27']= (S/len(pool_json_raw))
    sheet1['N27'].number_format='0.0000%'
    sheet1['O27']= (BD/len(pool_json_raw))
    sheet1['O27'].number_format='0.0000%'
    sheet1['P27']= (B/len(pool_json_raw))
    sheet1['P27'].number_format='0.0000%'
    sheet1['Q27']= (BL/len(pool_json_raw))
    sheet1['Q27'].number_format='0.0000%'
    sheet1['R27']= "=SUM(J27:Q27)"
    sheet1['R27'].number_format='0.0%'



    if len(pool_json_raw)==86400:
        workbook.save('{}\Analytics\Per_Detik\{} per {}.xlsx'.format(pool_path_simpan,nama_file_analitik,timeType))
    else: 
        workbook.save('{}\Analytics\Per_Menit\{} per {}.xlsx'.format(pool_path_simpan,nama_file_analitik,timeType))
