import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import filedialog, Text
import glob
import datetime
import os, psutil;
from method import *
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl.drawing.image import Image

from openpyxl.styles.borders import Border, Side

from openpyxl.styles import Font, Alignment,PatternFill

from openpyxl.utils import get_column_letter

def time_excel(time):
    date_time = datetime.datetime.strptime(time, "%H:%M:%S")
    a_timedelta = date_time - datetime.datetime(1900, 1, 1,)
    seconds = a_timedelta.total_seconds()
    # '{0:.5f}'.format(seconds/86400)
    output=seconds/86400
    return (output)


pool_json_raw=openFile()
    
file_wra = (glob.glob('{}\WRA\*.json'.format(pool_json_raw)))
file_wt = (glob.glob('{}\RTS\**\*.json'.format(pool_json_raw)))
pool_banyak_wt=os.listdir("{}\RTS".format(pool_json_raw))
counter_h=0
for h in file_wra:
    file_wra[counter_h]=os.path.normpath(file_wra[counter_h])
    counter_h+=1
counter_h=0
for h in file_wt:
    file_wt[counter_h]=os.path.normpath(file_wt[counter_h])
    counter_h+=1

for i in file_wra:
    file_name = os.path.basename(i).split('.')[0]
    file_name_custom=file_name.split('-')
    file_wra_yang_ini="{}-{}-{}".format(file_name_custom[1],file_name_custom[2],file_name_custom[3])
    retrieved_wt = list(filter(lambda x: file_wra_yang_ini in x, file_wt))
    print(file_wra_yang_ini)
    json_raw=panggil_json_raw(i)
    pool_wt=[]
    

    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Analytical"

    sheet1['A1']='Date'
    sheet1['B1']='Time'
    sheet1['C1']='WS'
    sheet1['D1']='Direction'
    sheet1['E1']='Degree'

    sheet1['F1']='U'
    sheet1['G1']='TL'
    sheet1['H1']='T'
    sheet1['I1']='Tg'
    sheet1['J1']='S'
    sheet1['K1']='BD'
    sheet1['L1']='B'
    sheet1['M1']='BL'

    mata_angin=['U','TL','T','Tg','S','BD','B','BL']
    


    iterate_row=2
    for k in json_raw:
        date_time = datetime.datetime.strptime(k['Time'], '%H:%M:%S').time()
        sheet1.cell(row=iterate_row, column=1).value = k['Date']
        sheet1.cell(row=iterate_row, column=2).value = date_time
        sheet1.cell(row=iterate_row, column=3).value = k['Kec_angin']
        sheet1.cell(row=iterate_row, column=4).value = k['Arah_angin']
        sheet1.cell(row=iterate_row, column=5).value = k['Derajat_angin']

        column_angin=6
        for p in mata_angin:
            # print(k['Arah_angin'])
            if (p == str(k['Arah_angin'])):
                sheet1.cell(row=iterate_row, column=column_angin).value = k['Kec_angin']
            else:
                sheet1.cell(row=iterate_row, column=column_angin).value = 0
            column_angin+=1
        
        if (k==json_raw[-1]):
            iterate_row=2
        else:
            iterate_row += 1
    range_wt=14+len(pool_banyak_wt)
    column_isi_wt=14
    # print(retrieved_wt)
    for r in retrieved_wt:
        json_wt=panggil_json_raw(r)
        # print(list(dict(json_wt[1]).values())[0])
        # print(r)
        for y in json_wt:
            # print(y)
            sheet1.cell(row=1, column=column_isi_wt).value = list(dict(json_wt[1]).values())[0]

            sheet1.cell(row=iterate_row, column=column_isi_wt).value = y['RTS']
                
            if (y==json_wt[-1]):
                # print(column_isi_wt)
                iterate_row=2
            else:
                iterate_row += 1
        column_isi_wt+=1
    # min = time_excel("09:20:00")
    # max = time_excel("09:35:00")
    # major=time_excel("00:03:00")

    chart1 = ScatterChart(scatterStyle="line")

    xvalues = Reference(sheet1, min_col = 2, min_row = 2, max_row = 86401)

    for e in range(6,14):
        values = Reference(sheet1, min_col = e,min_row = 1, max_row = 86401)
        series = Series(values = values, xvalues = xvalues,title_from_data=True)
        
        # chart1.x_axis.scaling.min = min
        # chart1.x_axis.scaling.max = max
        # chart1.x_axis.majorUnit = major
        chart1.series.append(series)
    chart1.y_axis.title = 'WS'

    chart2 = ScatterChart(scatterStyle="line")
    xvalues = Reference(sheet1, min_col = 2, min_row = 2, max_row = 86401)

    for w in range(14,column_isi_wt):
        values = Reference(sheet1, min_col = w,min_row = 1, max_row = 86401)
        series = Series(values = values, xvalues = xvalues,title_from_data=True)
        # chart2.x_axis.scaling.min = min
        # chart2.x_axis.scaling.max = max
        # chart2.x_axis.majorUnit = major
        chart2.series.append(series)
    chart2.y_axis.title = 'RPM'
    chart2.y_axis.axId = 200
    
    # chart1 += chart2
    # chart.width=19.5

    chart2.y_axis.crosses = "max"
    chart1 += chart2

    chart1.legend.position = 'b'
    chart1.width= 35
    chart1.height=17.5

    
    # add chart to the sheet
    # the top-left corner of a chart
    # is anchored to cell E2 .
    sheet1.add_chart(chart1, '{}{}'.format(get_column_letter(column_isi_wt),2))

            
    workbook.save('{}\RTS(A) {}.xlsx'.format(pool_json_raw,file_wra_yang_ini))

# print(file_wra)
# print(file_wt)
# print(pool_banyak_wt)